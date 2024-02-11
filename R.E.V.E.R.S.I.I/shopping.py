import speech_recognition as sr
from openpyxl import * #load_workbook
import pyttsx3


def shopping_list():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        Speak("what would you like to add?")
        print("What would you like to add to the shopping list?")
        audio = r.listen(source)
    data = ""
    try:
        data = r.recognize_google(audio)
        data.lower() 
        print("You said: " + data)
        wb=load_workbook("C:\\Users\\techmoto\\Desktop\\Voice\\shopping_list.xlsx")
        ws=wb["Sheet1"]
        mr = ws.max_row
        mc = ws.max_column
        mr = mr+1
        wcell1=ws.cell(mr,1)
        wcell1.value= data
        #wcell2=ws.cell(6,2)
        wb.save('C:\\Users\\techmoto\\Desktop\\Voice\\shopping_list.xlsx')
        Speak("I have added that to your shopping list")
    except sr.UnknownValueError:
        print("Jarvis did not understand your request")
    except sr.RequestError as e:
        print("Could not request results from Google Speech Recognition service; {0}".format(e))

def Speak(text):

    rate = 100 # setting the speech starting rate
    engine = pyttsx3.init()
    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[0].id) #0 is male and 1 is female
    engine.setProperty('rate', rate+50) #increasing speech rate by 50 (def = 200)
    engine.say(text)
    engine.runAndWait()

shopping_list()